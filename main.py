from fastapi import FastAPI, Depends, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import Employee, TimeEntry, get_db, init_db

ADMIN_PIN = "admin1234"

app = FastAPI(title="G&O Timesheet")
init_db()

app.mount("/static", StaticFiles(directory="static"), name="static")


# --- Schemas ---

class EmployeeCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    address: Optional[str] = None
    pin: str = "0000"


class LoginRequest(BaseModel):
    employee_id: int
    pin: str


class ClockInRequest(BaseModel):
    employee_id: int
    notes: Optional[str] = None


class ClockOutRequest(BaseModel):
    entry_id: int


# --- Helpers ---

def hours_between(clock_in: datetime, clock_out: datetime) -> float:
    return round((clock_out - clock_in).total_seconds() / 3600, 2)


def emp_dict(e: Employee):
    return {
        "id": e.id,
        "name": e.name,
        "email": e.email,
        "phone": e.phone,
        "address": e.address,
        "pin": e.pin,
        "created_at": e.created_at.isoformat() if e.created_at else None,
    }


# --- Routes ---

@app.get("/")
def root():
    return FileResponse("static/index.html")


# Self-register (public)
@app.post("/api/register", status_code=201)
def register(data: EmployeeCreate, db: Session = Depends(get_db)):
    if db.query(Employee).filter(Employee.email == data.email).first():
        raise HTTPException(status_code=400, detail="Email ya registrado / Email already registered")
    if not data.pin or len(data.pin) < 4:
        raise HTTPException(status_code=400, detail="PIN debe tener al menos 4 dígitos")
    emp = Employee(**data.model_dump())
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp_dict(emp)


# Auth
@app.post("/api/login")
def login(data: LoginRequest, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == data.employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    if emp.pin != data.pin:
        raise HTTPException(status_code=401, detail="PIN incorrecto")
    return emp_dict(emp)


@app.post("/api/admin-login")
def admin_login(data: dict):
    if data.get("pin") != ADMIN_PIN:
        raise HTTPException(status_code=401, detail="PIN de administrador incorrecto")
    return {"ok": True}


# Employees (admin)
@app.get("/api/employees")
def list_employees(db: Session = Depends(get_db)):
    return [emp_dict(e) for e in db.query(Employee).order_by(Employee.name).all()]


@app.post("/api/employees", status_code=201)
def create_employee(data: EmployeeCreate, db: Session = Depends(get_db)):
    if db.query(Employee).filter(Employee.email == data.email).first():
        raise HTTPException(status_code=400, detail="Email ya registrado")
    emp = Employee(**data.model_dump())
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp_dict(emp)


@app.get("/api/employees/{emp_id}")
def get_employee(emp_id: int, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == emp_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    return emp_dict(emp)


@app.put("/api/employees/{emp_id}")
def update_employee(emp_id: int, data: EmployeeCreate, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == emp_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    for k, v in data.model_dump().items():
        setattr(emp, k, v)
    db.commit()
    db.refresh(emp)
    return emp_dict(emp)


@app.delete("/api/employees/{emp_id}")
def delete_employee(emp_id: int, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == emp_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    db.query(TimeEntry).filter(TimeEntry.employee_id == emp_id).delete()
    db.delete(emp)
    db.commit()
    return {"ok": True}


# Clock in/out
@app.post("/api/clock-in")
def clock_in(data: ClockInRequest, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == data.employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    if db.query(TimeEntry).filter(
        TimeEntry.employee_id == data.employee_id, TimeEntry.clock_out == None
    ).first():
        raise HTTPException(status_code=400, detail="Ya tienes una entrada activa")
    entry = TimeEntry(employee_id=data.employee_id, clock_in=datetime.utcnow(), notes=data.notes)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "clock_in": entry.clock_in.isoformat(), "notes": entry.notes}


@app.post("/api/clock-out")
def clock_out(data: ClockOutRequest, db: Session = Depends(get_db)):
    entry = db.query(TimeEntry).filter(TimeEntry.id == data.entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")
    if entry.clock_out:
        raise HTTPException(status_code=400, detail="Ya se registró la salida")
    entry.clock_out = datetime.utcnow()
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "clock_out": entry.clock_out.isoformat()}


@app.get("/api/my-status/{emp_id}")
def my_status(emp_id: int, db: Session = Depends(get_db)):
    open_entry = db.query(TimeEntry).filter(
        TimeEntry.employee_id == emp_id, TimeEntry.clock_out == None
    ).first()
    if open_entry:
        return {"active": True, "entry_id": open_entry.id, "clock_in": open_entry.clock_in.isoformat()}
    return {"active": False}


@app.get("/api/my-history/{emp_id}")
def my_history(emp_id: int, db: Session = Depends(get_db)):
    entries = db.query(TimeEntry).filter(
        TimeEntry.employee_id == emp_id,
        TimeEntry.clock_out != None
    ).order_by(TimeEntry.clock_in.desc()).limit(50).all()
    return [
        {
            "id": e.id,
            "clock_in": e.clock_in.isoformat(),
            "clock_out": e.clock_out.isoformat(),
            "hours": hours_between(e.clock_in, e.clock_out),
            "notes": e.notes or "",
        }
        for e in entries
    ]


# Admin reports
@app.get("/api/report")
def report(
    employee_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(TimeEntry).filter(TimeEntry.clock_out != None)
    if employee_id:
        query = query.filter(TimeEntry.employee_id == employee_id)
    if date_from:
        query = query.filter(TimeEntry.clock_in >= datetime.fromisoformat(date_from))
    if date_to:
        dt_to = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59)
        query = query.filter(TimeEntry.clock_in <= dt_to)
    entries = query.order_by(TimeEntry.clock_in.desc()).all()
    result = []
    for e in entries:
        emp = db.query(Employee).filter(Employee.id == e.employee_id).first()
        result.append({
            "entry_id": e.id,
            "employee_id": e.employee_id,
            "employee_name": emp.name if emp else "—",
            "clock_in": e.clock_in.isoformat(),
            "clock_out": e.clock_out.isoformat(),
            "hours": hours_between(e.clock_in, e.clock_out),
            "notes": e.notes or "",
        })
    return result


@app.get("/api/export")
def export_excel(
    employee_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    data = report(employee_id=employee_id, date_from=date_from, date_to=date_to, db=db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    hf = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    headers = ["Empleado", "Entrada", "Salida", "Horas", "Notas"]
    widths = [25, 22, 22, 10, 30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w
    for row, e in enumerate(data, 2):
        ws.cell(row=row, column=1, value=e["employee_name"])
        ws.cell(row=row, column=2, value=e["clock_in"].replace("T", " ").split(".")[0])
        ws.cell(row=row, column=3, value=e["clock_out"].replace("T", " ").split(".")[0])
        ws.cell(row=row, column=4, value=e["hours"])
        ws.cell(row=row, column=5, value=e["notes"])
    last = len(data) + 2
    ws.cell(row=last, column=3, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=last, column=4, value=round(sum(e["hours"] for e in data), 2)).font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=timesheet.xlsx"}
    )
